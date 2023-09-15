# Detailed Description
# This Python script fetches a list of movies from The Movie Database (TMDB) API based on multiple search criteria.
# The search criteria include release year, genre, keyword, company, and persons.
# The script fetches the data, converts it into a readable format, and exports it to an Excel file.
# Additionally, it exports the TMDB IDs of the fetched movies to a separate CSV file.

import requests
import openpyxl
import os
import csv
from datetime import datetime

# Function to export TMDB IDs to a separate CSV file
def export_to_csv(movie_list, csv_path):
    with open(csv_path, 'w', newline='') as csvfile:
        csv_writer = csv.writer(csvfile)
        csv_writer.writerow(['tmdb'])
        for movie in movie_list:
            csv_writer.writerow([movie[2]])

# Function to fetch movies based on different criteria
def fetch_movies(api_key, release_year="", genres="", keywords="", companies="", persons=""):
    base_url = "https://api.themoviedb.org/3/discover/movie"
    params = {
        "api_key": api_key,
        "sort_by": "popularity.desc"
    }
    if release_year:
        params["primary_release_year"] = release_year
    if genres:
        params["with_genres"] = genres
    if keywords:
        params["with_keywords"] = keywords
    if companies:
        params["with_companies"] = companies
    if persons:
        params["with_people"] = persons

    movie_list = []
    current_page = 1

    print("\nFetching movies...")

    while True:
        params["page"] = current_page
        response = requests.get(base_url, params=params)

        if response.status_code == 200:
            data = response.json()
            results = data.get("results")

            if results:
                for movie in results:
                    title = movie.get('title')
                    movie_id = movie.get('id')
                    release_year_movie = movie.get('release_date')[:4]
                    movie_list.append((title, release_year_movie, movie_id))
                current_page += 1
            else:
                break
        else:
            print(f"\nError occurred while fetching data from TMDB for Page {current_page}.")
            break

    return movie_list

# Function to convert genre names to IDs
def genre_name_to_id(api_key, genre_names):
    genre_list_url = f"https://api.themoviedb.org/3/genre/movie/list?api_key={api_key}&language=en-US"
    response = requests.get(genre_list_url)
    if response.status_code == 200:
        data = response.json()
        genres = data.get("genres")
        genre_map = {genre['name'].lower(): genre['id'] for genre in genres}
        genre_ids = [str(genre_map[name.lower()]) for name in genre_names if name.lower() in genre_map]
        return ",".join(genre_ids)
    else:
        print("Failed to fetch genre list.")
        return ""

# Function to convert person names to IDs
def person_name_to_id(api_key, person_names):
    person_ids = []
    for person_name in person_names:
        search_url = f"https://api.themoviedb.org/3/search/person?api_key={api_key}&query={person_name}"
        response = requests.get(search_url)
        if response.status_code == 200:
            data = response.json()
            results = data.get("results")
            if results:
                person_ids.append(str(results[0]['id']))
    return ",".join(person_ids)

# Function to convert company names to IDs
def company_name_to_id(api_key, company_names):
    company_ids = []
    for company_name in company_names:
        search_url = f"https://api.themoviedb.org/3/search/company?api_key={api_key}&query={company_name}"
        response = requests.get(search_url)
        if response.status_code == 200:
            data = response.json()
            results = data.get("results")
            if results:
                company_ids.append(str(results[0]['id']))
    return ",".join(company_ids)

# Main function
def main():
    api_key = "blah blah fuck off"

    release_year = input("Enter the desired release year (or leave blank for all years): ")
    genre_names = input("Enter the genre names separated by commas (or leave blank for all genres): ").split(",")
    keyword_ids = input("Enter the keyword IDs separated by commas (or leave blank for all keywords): ").split(",")
    company_names = input("Enter the company names separated by commas (or leave blank for all companies): ").split(",")
    person_names = input("Enter the person names separated by commas (or leave blank for all persons): ").split(",")

    if genre_names[0]:
        genre_ids = genre_name_to_id(api_key, genre_names)
    else:
        genre_ids = ""

    if company_names[0]:
        company_ids = company_name_to_id(api_key, company_names)
    else:
        company_ids = ""

    if person_names[0]:
        person_ids = person_name_to_id(api_key, person_names)
    else:
        person_ids = ""

    movie_list = fetch_movies(api_key, release_year, genre_ids, ",".join(keyword_ids), company_ids, person_ids)

    current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    excel_file_path = f"C:\\Users\\bax11\\OneDrive\\Desktop\\Script_Tasks\\Excel_Movie_Lists\\TMDB_{release_year}_as_of_{current_time}.xlsx"

    if os.path.exists(excel_file_path):
        wb = openpyxl.load_workbook(excel_file_path)
    else:
        wb = openpyxl.Workbook()

    ws = wb.active

    if ws.max_row == 1:
        ws.append(["Movie Title", "Release Year", "TMDB Movie ID"])

    for movie_details in movie_list:
        ws.append(movie_details)

    wb.save(excel_file_path)
    print(f"\nMovies details have been exported to {excel_file_path}.")

    # Export to CSV
    csv_path = "C:\\Users\\bax11\\OneDrive\\Desktop\\radarr_scripts\\trakt_import\\trakt\\scrapped.csv"
    export_to_csv(movie_list, csv_path)

if __name__ == '__main__':
    main()
